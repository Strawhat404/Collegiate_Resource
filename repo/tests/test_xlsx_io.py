"""Excel bulk import / export round-trip."""
from __future__ import annotations
try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


def _write_xlsx(path, header, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    wb.save(str(path))


def test_xlsx_import_and_export(container, admin_session, tmp_path):
    if openpyxl is None:
        return  # environment lacks openpyxl
    src = tmp_path / "in.xlsx"
    _write_xlsx(src,
                ["student_id", "full_name", "college", "class_year",
                 "email", "phone", "housing_status"],
                [["X100", "Xena", "LA", 2026, "x@e.edu", "555-1212", "pending"],
                 ["X101", "Yan",  "EN", 2025, "y@e.edu", "555-1213", "pending"]])
    preview = container.students.import_file(admin_session, src)
    assert len(preview.accepted) == 2
    assert preview.rejected == []
    container.students.commit_import(admin_session, preview.preview_id)
    out = tmp_path / "out.xlsx"
    n = container.students.export_file(admin_session, out)
    assert n >= 2 and out.is_file()
    wb = openpyxl.load_workbook(str(out))
    ws = wb.active
    assert ws.max_row >= 3  # header + 2 rows
